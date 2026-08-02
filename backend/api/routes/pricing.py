"""Engineering Pricing (工程计价) CRUD routes."""
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Body
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from core.database import get_db
from core.security import get_current_user
from models import EngineeringPricing, Project, AuditLog, User
from schemas import EngineeringPricingCreate, EngineeringPricingResponse

router = APIRouter(prefix="/pricing", tags=["工程计价"])


def _is_admin(user: User) -> bool:
    return user.role in ("super_admin", "company_admin")


def _can_create(user: User) -> bool:
    """可录入的角色：管理员 + 考勤员"""
    return user.role in ("super_admin", "company_admin", "attendance")


@router.post("", response_model=EngineeringPricingResponse)
async def create_pricing(
    req: EngineeringPricingCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """创建工程计价记录 — 管理员和考勤员可录入"""
    if not _can_create(user):
        raise HTTPException(status_code=403, detail="权限不足")
    item = EngineeringPricing(tenant_id=user.tenant_id or 1, **req.model_dump())
    db.add(item)
    await db.flush()
    await db.refresh(item)
    db.add(AuditLog(tenant_id=user.tenant_id or 1, user_id=user.id, username=user.username,
                    action="create_pricing", biz_type="pricing", biz_id=item.id))
    return item


@router.get("")
async def list_pricing(
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    page: int = 1,
    page_size: int = 10,
):
    """工程计价列表 — 项目负责人只读，考勤员/管理员可查看"""
    page_size = min(page_size, 100)
    query = select(EngineeringPricing)
    count_q = select(func.count(EngineeringPricing.id))
    if user.role != "super_admin" and user.tenant_id:
        query = query.where(EngineeringPricing.tenant_id == user.tenant_id)
        count_q = count_q.where(EngineeringPricing.tenant_id == user.tenant_id)
    if project_id:
        query = query.where(EngineeringPricing.project_id == project_id)
        count_q = count_q.where(EngineeringPricing.project_id == project_id)
    if category:
        query = query.where(EngineeringPricing.category == category)
        count_q = count_q.where(EngineeringPricing.category == category)
    total = (await db.execute(count_q)).scalar() or 0
    result = await db.execute(
        query.order_by(EngineeringPricing.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    return {"items": result.scalars().all(), "total": total, "page": page, "page_size": page_size}


# ─── 批量审核（必须在 /{item_id} 之前定义） ───────────────────

@router.put("/batch-approve")
async def batch_approve_pricing(
    ids: List[int] = Body(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量审核工程计价记录 — 仅管理员，只审核未审核的"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="权限不足")
    if not ids:
        raise HTTPException(status_code=400, detail="未选择记录")
    result = await db.execute(
        select(EngineeringPricing).where(
            EngineeringPricing.id.in_(ids),
            EngineeringPricing.is_approved == False,
        )
    )
    items = result.scalars().all()
    for item in items:
        item.is_approved = True
    await db.flush()
    return {"detail": f"已批量审核 {len(items)} 条记录"}


@router.put("/{item_id}", response_model=EngineeringPricingResponse)
async def update_pricing(
    item_id: int,
    req: EngineeringPricingCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """编辑工程计价记录 — 管理员可编辑未审核记录，已审核不可编辑"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="权限不足")
    result = await db.execute(select(EngineeringPricing).where(EngineeringPricing.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    if item.is_approved:
        raise HTTPException(status_code=403, detail="已审核的记录不可编辑")
    for key, val in req.model_dump().items():
        setattr(item, key, val)
    await db.flush()
    await db.refresh(item)
    return item


@router.delete("/{item_id}")
async def delete_pricing(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """删除工程计价记录 — 管理员可删除未审核记录，已审核不可删除"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="权限不足")
    result = await db.execute(select(EngineeringPricing).where(EngineeringPricing.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    if item.is_approved:
        raise HTTPException(status_code=403, detail="已审核的记录不可删除")
    await db.delete(item)
    await db.flush()
    return {"detail": "删除成功"}


@router.put("/{item_id}/approve", response_model=EngineeringPricingResponse)
async def approve_pricing(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """审核工程计价记录 — 管理员可切换审核状态"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="权限不足")
    result = await db.execute(select(EngineeringPricing).where(EngineeringPricing.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    item.is_approved = not item.is_approved
    await db.flush()
    await db.refresh(item)
    return item


@router.get("/total")
async def pricing_total(
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """工程计价汇总金额"""
    q = select(func.coalesce(func.sum(EngineeringPricing.amount), 0))
    if user.role != "super_admin" and user.tenant_id:
        q = q.where(EngineeringPricing.tenant_id == user.tenant_id)
    if project_id:
        q = q.where(EngineeringPricing.project_id == project_id)
    if category:
        q = q.where(EngineeringPricing.category == category)
    total = (await db.execute(q)).scalar()
    return {"total": float(total)}


@router.get("/export")
async def export_pricing(
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出工程计价为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    query = select(EngineeringPricing)
    if user.role != "super_admin" and user.tenant_id:
        query = query.where(EngineeringPricing.tenant_id == user.tenant_id)
    if project_id:
        query = query.where(EngineeringPricing.project_id == project_id)
    if category:
        query = query.where(EngineeringPricing.category == category)
    result = await db.execute(query.order_by(EngineeringPricing.id.desc()))
    items = result.scalars().all()

    # Resolve project names
    pids = {p.project_id for p in items}
    projects_q = await db.execute(select(Project).where(Project.id.in_(list(pids) or [0])))
    pname = {pp.id: pp.name for pp in projects_q.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工程计价"

    h_font = Font(bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["项目名称", "单项工程名称", "金额", "日期", "审核状态", "备注"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_align; c.border = thin

    for ri, item in enumerate(items, 2):
        vals = [
            pname.get(item.project_id, ""),
            item.item_name,
            float(item.amount or 0),
            str(item.pricing_date or ""),
            "已审核" if item.is_approved else "待审核",
            item.remark or "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = h_align; c.border = thin

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=pricing.xlsx"})
