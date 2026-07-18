"""Project (项目) CRUD."""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select, func
from sqlalchemy.orm import joinedload
from sqlalchemy.ext.asyncio import AsyncSession
from core.database import get_db
from core.security import get_current_user, require_project
from models import Project, Partner, AuditLog, User
from schemas import ProjectCreate, ProjectResponse

router = APIRouter(prefix="/projects", tags=["项目管理"])


@router.post("", response_model=ProjectResponse)
async def create_project(
    req: ProjectCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
):
    """创建项目 — 管理员和项目负责人可操作"""
    data = req.model_dump()
    # Auto-calculate subcontract ratio
    total = (data.get("labor_subcontract_amount") or 0) + \
            (data.get("machinery_rental_amount") or 0) + \
            (data.get("live_working_amount") or 0)
    contract = data.get("contract_amount") or 0
    data["subcontract_ratio"] = round(float(total) / float(contract), 4) if contract else None
    project = Project(**data)
    tid = user.tenant_id or 1
    project.tenant_id = tid
    db.add(project)
    await db.flush()
    db.add(AuditLog(tenant_id=tid, user_id=user.id, username=user.username,
                    action="create_project", biz_type="project", biz_id=project.id))
    # Reload with owner name
    result = await db.execute(
        select(Project).where(Project.id == project.id)
    )
    return result.scalar_one()


@router.get("")
async def list_projects(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
):
    page_size = min(page_size, 100)
    query = select(Project)
    count_q = select(func.count(Project.id))
    if user.role != "super_admin" and user.tenant_id:
        query = query.where(Project.tenant_id == user.tenant_id)
        count_q = count_q.where(Project.tenant_id == user.tenant_id)
    if search:
        query = query.where(Project.name.ilike(f"%{search}%"))
        count_q = count_q.where(Project.name.ilike(f"%{search}%"))
    total = (await db.execute(count_q)).scalar() or 0
    result = await db.execute(query.order_by(Project.id.desc()).offset((page - 1) * page_size).limit(page_size))
    projects = result.scalars().all()

    # Compute revenue and unpaid amounts from invoice data
    from sqlalchemy import func
    from models import InvoiceOut

    # Batch-load all invoice revenue for these projects
    pids = [p.id for p in projects] or [0]
    rev_q = await db.execute(
        select(InvoiceOut.project_id, func.coalesce(func.sum(InvoiceOut.amount), 0).label("revenue"))
        .where(InvoiceOut.project_id.in_(pids))
        .group_by(InvoiceOut.project_id)
    )
    revenue_map = {row.project_id: float(row.revenue) for row in rev_q.all()}

    responses = []
    for p in projects:
        owner_name = ""
        if p.owner_id:
            r = await db.execute(select(Partner.name).where(Partner.id == p.owner_id))
            if name := r.scalar_one_or_none():
                owner_name = name
        resp = ProjectResponse.model_validate(p)
        resp.owner_name = owner_name
        resp.revenue_amount = revenue_map.get(p.id, 0.0)
        settlement = float(p.settlement_amount or 0)
        resp.unpaid_amount = round(settlement - resp.revenue_amount, 2)
        responses.append(resp)
    return {"items": responses, "total": total, "page": page, "page_size": page_size}


@router.get("/export")
async def export_projects(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
    search: Optional[str] = None,
):
    """导出项目列表为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    query = select(Project)
    if user.role != "super_admin" and user.tenant_id:
        query = query.where(Project.tenant_id == user.tenant_id)
    if search:
        query = query.where(Project.name.ilike(f"%{search}%"))
    result = await db.execute(query.order_by(Project.id.desc()))
    projects = result.scalars().all()

    # Compute revenue
    from sqlalchemy import func
    from models import InvoiceOut
    pids = [p.id for p in projects] or [0]
    rev_q = await db.execute(
        select(InvoiceOut.project_id, func.coalesce(func.sum(InvoiceOut.amount), 0).label("revenue"))
        .where(InvoiceOut.project_id.in_(pids)).group_by(InvoiceOut.project_id)
    )
    revenue_map = {row.project_id: float(row.revenue) for row in rev_q.all()}

    # Resolve partner names
    partner_ids = set()
    for p in projects:
        if p.owner_id: partner_ids.add(p.owner_id)
        if p.winning_bid_unit_id: partner_ids.add(p.winning_bid_unit_id)
    partners_q = await db.execute(select(Partner).where(Partner.id.in_(list(partner_ids) or [0])))
    pname = {pp.id: pp.name for pp in partners_q.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目管理"

    # Header style
    h_font = Font(bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["项目名称", "业主", "中标单位", "合同金额", "审定金额", "收入金额", "未付金额",
               "劳务分包", "机械租赁", "带电作业", "分包比例", "签订日期", "计划开工", "计划竣工",
               "实际开工", "实际竣工", "状态"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_align; c.border = thin

    for ri, p in enumerate(projects, 2):
        rev = revenue_map.get(p.id, 0.0)
        sett = float(p.settlement_amount or 0)
        unpaid = round(sett - rev, 2)
        total_sub = (float(p.labor_subcontract_amount or 0) + float(p.machinery_rental_amount or 0) + float(p.live_working_amount or 0))
        ratio = round(total_sub / float(p.contract_amount or 1), 4) if p.contract_amount else None

        vals = [
            p.name,
            pname.get(p.owner_id, ""),
            pname.get(p.winning_bid_unit_id, ""),
            float(p.contract_amount or 0),
            sett, rev, unpaid,
            float(p.labor_subcontract_amount or 0),
            float(p.machinery_rental_amount or 0),
            float(p.live_working_amount or 0),
            f"{ratio*100:.1f}%" if ratio else "-",
            p.contract_sign_date or "",
            p.start_date or "", p.end_date or "",
            p.actual_start_date or "", p.actual_end_date or "",
            p.status,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = h_align; c.border = thin

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=projects.xlsx"})


@router.get("/{project_id}", response_model=ProjectResponse)
async def get_project(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
):
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return project


@router.put("/{project_id}", response_model=ProjectResponse)
async def update_project(
    project_id: int,
    req: ProjectCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
):
    """编辑项目 — 管理员和项目负责人可操作"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    data = req.model_dump()
    # Auto-calculate subcontract ratio
    total = (data.get("labor_subcontract_amount") or 0) + \
            (data.get("machinery_rental_amount") or 0) + \
            (data.get("live_working_amount") or 0)
    contract = data.get("contract_amount") or 0
    data["subcontract_ratio"] = round(float(total) / float(contract), 4) if contract else None
    for key, val in data.items():
        setattr(project, key, val)
    await db.flush()
    await db.refresh(project)
    return project


@router.delete("/{project_id}")
async def delete_project(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
):
    """删除项目 — 管理员和项目负责人可操作"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    tid = user.tenant_id or 1
    await db.delete(project)
    await db.flush()
    db.add(AuditLog(tenant_id=tid, user_id=user.id, username=user.username,
                    action="delete_project", biz_type="project", biz_id=project_id))
    return {"detail": "删除成功"}


@router.get("/{project_id}/profit")
async def get_project_profit(
    project_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_project),
):
    """单项目利润核算"""
    from sqlalchemy import func
    from models import InvoiceOut, InvoiceIn, CostRecord

    # 总收入（销项发票）
    r1 = await db.execute(
        select(func.coalesce(func.sum(InvoiceOut.amount), 0))
        .where(InvoiceOut.project_id == project_id)
    )
    total_income = r1.scalar()

    # 总成本（进项发票 + 人工 + 杂费）
    r2 = await db.execute(
        select(func.coalesce(func.sum(InvoiceIn.amount), 0))
        .where(InvoiceIn.project_id == project_id)
    )
    total_invoice_cost = r2.scalar()

    r3 = await db.execute(
        select(func.coalesce(func.sum(CostRecord.amount), 0))
        .where(CostRecord.project_id == project_id)
    )
    total_other_cost = r3.scalar()

    profit = total_income - total_invoice_cost - total_other_cost

    return {
        "project_id": project_id,
        "total_income": float(total_income),
        "total_invoice_cost": float(total_invoice_cost),
        "total_other_cost": float(total_other_cost),
        "profit": float(profit),
    }
